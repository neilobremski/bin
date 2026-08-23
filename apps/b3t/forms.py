"""Microsoft Forms submission download and parsing.

Deterministic flow:
    1. Navigate to Forms responses page (&analysis=true)
    2. Dismiss "Got it" popup if present
    3. Click "More options" next to Excel
    4. Click "Download a copy"
    5. Move downloaded .xlsx to edition wip/
"""
import json
import os
import shutil
import sys
import time
from urllib.parse import urlparse

import session
import env
from constants import FORMS_URL, FORMS_DOWNLOAD_PREFIX

# Forms responses page with analysis=true to land on the Excel view
FORMS_RESPONSES_URL = FORMS_URL + "&analysis=true"


def dispatch(args):
    action = args.action
    if not action:
        print("Usage: b3t forms <login|download|list>", file=sys.stderr)
        return 2
    if action == "login":
        return cmd_login(args)
    elif action == "download":
        return cmd_download(args)
    elif action == "list":
        return cmd_list(args)
    return 2


FORMS_HOSTS = ("forms.office.com", "forms.microsoft.com", "forms.cloud.microsoft")


def _is_forms_url(url):
    hostname = (urlparse(url).hostname or "").lower()
    return any(hostname == host or hostname.endswith(f".{host}") for host in FORMS_HOSTS)


def ensure_authenticated():
    """Check M365 auth by navigating to Forms. Returns True if authed."""
    session.ensure_running()

    session.navigate("https://forms.office.com")
    # Allow time for the SSO redirect chain to settle; login.microsoftonline.com
    # mid-chain is not a failure until it stops moving.
    url = None
    for _ in range(4):
        time.sleep(3)
        url = session.current_url()
        if url and _is_forms_url(url):
            return True

    print(f"ERROR: Not authenticated to M365 (stuck at: {url}).", file=sys.stderr)
    print("Run: b3t open, log into forms.office.com, b3t close", file=sys.stderr)
    return False


def cmd_login(args):
    if ensure_authenticated():
        print("M365 authenticated.", file=sys.stderr)
        return 0
    return 1


def cmd_download(args):
    """Download submissions Excel — deterministic click sequence."""
    import re

    if not ensure_authenticated():
        return 1

    edition_dir = os.path.join(os.getcwd(), "editions", args.edition, "wip")
    os.makedirs(edition_dir, exist_ok=True)

    # Step 1: Navigate to Forms responses/analysis page
    print("Navigating to Forms responses...", file=sys.stderr)
    session.navigate(FORMS_RESPONSES_URL)
    time.sleep(4)

    # Step 2: Dismiss "Got it" popup if present
    snap = session.snapshot()
    if snap:
        ref = _find_ref(snap, lambda l: "Got it" in l and "button" in l.lower())
        if ref:
            session.run("click", ref)
            time.sleep(1)

    # Step 3: Click "More options" button (near the Excel section, not the Responses one)
    snap = session.snapshot()
    if not snap:
        print("ERROR: Cannot get page snapshot.", file=sys.stderr)
        return 1

    # Find the "More options" button that's NOT "More options for Responses"
    more_ref = None
    for line in snap.split("\n"):
        if 'button "More options"' in line and "Responses" not in line:
            m = re.search(r'\[ref=(\w+)\]', line)
            if m:
                more_ref = m.group(1)
                break
    if not more_ref:
        # Fallback: any "More options" button
        more_ref = _find_ref(snap, lambda l: "More options" in l and "button" in l.lower() and "Responses" not in l)

    if not more_ref:
        print("ERROR: Cannot find 'More options' button.", file=sys.stderr)
        return 1

    session.run("click", more_ref)
    time.sleep(1)

    # Step 4: Click "Download a copy" menuitem
    snap = session.snapshot()
    dl_ref = _find_ref(snap, lambda l: "Download a copy" in l and "menuitem" in l.lower())
    if not dl_ref:
        print("ERROR: Cannot find 'Download a copy' menu item.", file=sys.stderr)
        return 1

    download_start = time.time()
    session.run("click", dl_ref)
    print("Downloading...", file=sys.stderr)

    # Step 5: Find downloaded file. CDP-attached real Chrome saves to
    # ~/Downloads; the playwright-cli managed browser saves to .playwright-cli/.
    dl_dirs = [
        os.path.join(os.getcwd(), ".playwright-cli"),
        os.path.expanduser("~/Downloads"),
    ]
    deadline = time.time() + 30
    xlsx_files = []
    while time.time() < deadline:
        for d in dl_dirs:
            if not os.path.isdir(d):
                continue
            candidates = [f for f in os.listdir(d) if f.endswith(".xlsx")]
            if FORMS_DOWNLOAD_PREFIX:
                candidates = [f for f in candidates if FORMS_DOWNLOAD_PREFIX in f]
            candidates = [
                os.path.join(d, f)
                for f in candidates
                if os.path.getmtime(os.path.join(d, f)) >= download_start - 2
            ]
            xlsx_files.extend(candidates)
        if xlsx_files:
            break
        time.sleep(1)

    xlsx_files.sort(key=os.path.getmtime, reverse=True)

    if not xlsx_files:
        print("ERROR: No .xlsx downloaded.", file=sys.stderr)
        return 1

    src = xlsx_files[0]
    dst = os.path.join(edition_dir, "submissions.xlsx")
    shutil.copy2(src, dst)
    os.remove(src)
    print(f"Saved: {dst}", file=sys.stderr)
    print(dst)
    return 0


def _find_ref(snapshot_text, test_fn):
    """Find first ref matching test_fn(line) -> bool."""
    import re
    for line in snapshot_text.split("\n"):
        if test_fn(line):
            m = re.search(r'\[ref=(\w+)\]', line)
            if m:
                return m.group(1)
    return None


def cmd_list(args):
    """Parse submissions from downloaded Excel."""
    from datetime import date, datetime

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed.", file=sys.stderr)
        return 1

    xlsx_path = os.path.join(os.getcwd(), "editions", args.edition, "wip", "submissions.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"ERROR: {xlsx_path} not found. Run: b3t forms download --edition {args.edition}", file=sys.stderr)
        return 1

    since = None
    if args.since:
        since = datetime.fromisoformat(args.since)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    time_col = "Completion time" if "Completion time" in headers else "Start time"
    if time_col not in headers:
        print(f"ERROR: Expected '{time_col}' column in {xlsx_path}. Found: {headers}", file=sys.stderr)
        wb.close()
        return 1

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if not record.get("Title") and not record.get("Article Body"):
            continue
        submitted = record.get(time_col)
        if submitted is None:
            continue
        if isinstance(submitted, str):
            try:
                submitted = datetime.fromisoformat(submitted)
            except ValueError:
                continue
        if since and submitted < since:
            continue
        rows.append(record)

    wb.close()

    if hasattr(args, "json") and args.json:
        def default(o):
            if hasattr(o, "isoformat"):
                return o.isoformat()
            return str(o)
        json.dump(rows, sys.stdout, indent=2, default=default)
        print()
    else:
        print(f"{len(rows)} submissions", file=sys.stderr)
        for r in rows:
            title = r.get("Title") or "?"
            name = r.get("Name2") or r.get("Name") or "?"
            cat = r.get("Article Category") or ""
            print(f"  {name}: {title} [{cat}]")

    return 0
